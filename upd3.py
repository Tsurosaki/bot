import urllib.parse
import os
import logging
import asyncio
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Tuple
from collections import defaultdict

from aiogram import Bot, Dispatcher, types, F
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.filters import Command, StateFilter
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, ContentType, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, FSInputFile
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram import Router
from aiogram.utils.media_group import MediaGroupBuilder

import yadisk
import pandas as pd
from openpyxl import Workbook, load_workbook

# ========== НАСТРОЙКИ ==========
TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN')
YANDEX_TOKEN = os.getenv('YANDEX_TOKEN')
BASE_YANDEX_FOLDER = os.getenv('BASE_YANDEX_FOLDER', '/Фото за МК')
BASE_FOLDER_URL = os.getenv('BASE_FOLDER_URL', 'https://yadi.sk/d/xq_U3H4ygvkLiw')

# Настройки для Excel
EXCEL_FILE = "bot_data.xlsx"
EXCEL_SHEET = "Messages"

# Настройки для бэкапов
BACKUP_FOLDER = "backups"
BACKUP_INTERVAL_HOURS = 24  # Интервал между бэкапами в часах
BACKUP_RETENTION_DAYS = 30  # Хранить бэкапы 30 дней

# URL с гайдом (можно загрузить на Google Drive, Яндекс.Диск и т.д.)
GUIDE_URL = "https://disk.yandex.ru/i/1nzxVEriVKKYIg"  # Замените на реальную ссылку

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ========== STATES ==========
class PhotoStates(StatesGroup):
    waiting_for_folder_name = State()
    choosing_folder = State()
    waiting_for_custom_name = State()
    processing_album = State()
    waiting_for_batch_choice = State()
    waiting_for_folder_selection = State()

class TextStates(StatesGroup):
    waiting_for_fio = State()  # Ожидание ввода ФИО
    waiting_for_text = State()  # Ожидание текста для сохранения
    waiting_for_text_description = State()  # Ожидание описания для текста
    skip_text_confirmation = State()  # Подтверждение пропуска текста

class ExportStates(StatesGroup):
    waiting_for_user_selection = State()  # Ожидание выбора пользователя для экспорта

# ========== EXCEL MANAGER ==========
class ExcelManager:
    def __init__(self, filename: str = EXCEL_FILE):
        self.filename = filename
        self._init_excel()
    
    def _init_excel(self):
        """Инициализация Excel файла если его нет"""
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = EXCEL_SHEET
            
            # Создаем заголовки
            headers = [
                "ID", "Дата и время", "User ID", "Username", 
                "ФИО", "Текст", "Описание", "Тип"
            ]
            ws.append(headers)
            
            # Устанавливаем ширину колонок
            for i, header in enumerate(headers, 1):
                ws.column_dimensions[chr(64 + i)].width = 20
            
            wb.save(self.filename)
            logger.info(f"Создан новый Excel файл: {self.filename}")
    
    def _create_backup(self, backup_type: str = "manual") -> Optional[str]:
        """Создает бэкап файла перед изменениями"""
        try:
            if not os.path.exists(self.filename):
                return None
            
            # Создаем папку для бэкапов если её нет
            Path(BACKUP_FOLDER).mkdir(exist_ok=True)
            
            # Формируем имя бэкапа
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{backup_type}_backup_{timestamp}.xlsx"
            backup_path = os.path.join(BACKUP_FOLDER, backup_name)
            
            # Копируем файл
            import shutil
            shutil.copy2(self.filename, backup_path)
            logger.info(f"Создан бэкап: {backup_path}")
            
            return backup_path
            
        except Exception as e:
            logger.error(f"Ошибка при создании бэкапа: {e}")
            return None
    
    def save_message(self, user_id: int, fio: str, text: str, description: str = "", msg_type: str = "text"):
        """Сохраняет сообщение в Excel"""
        try:
            # Создаем бэкап перед изменением
            self._create_backup("auto")
            
            # Загружаем существующий файл
            wb = load_workbook(self.filename)
            ws = wb[EXCEL_SHEET]
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Получаем следующий ID (максимальный ID + 1)
            max_id = 0
            for row in range(2, ws.max_row + 1):  # начинаем с 2, пропуская заголовок
                cell_value = ws.cell(row, 1).value  # колонка A - ID
                if cell_value and isinstance(cell_value, (int, float)):
                    max_id = max(max_id, int(cell_value))
            
            next_id = max_id + 1
            
            # Получаем username из базы или используем пустую строку
            username = ""  # Будет обновлено позже через update_user_info
            
            # Добавляем новую строку
            ws.append([
                next_id,
                now,
                user_id,
                username,
                fio,
                text if text else "Без текста",
                description if description else "Без описания",
                msg_type
            ])
            
            wb.save(self.filename)
            logger.info(f"Сообщение сохранено в Excel. User ID: {user_id}, ФИО: {fio}, ID: {next_id}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении в Excel: {e}")
            return False
    
    def update_user_info(self, user_id: int, username: str = None, fio: str = None):
        """Обновляет информацию о пользователе во всех его записях"""
        try:
            wb = load_workbook(self.filename)
            ws = wb[EXCEL_SHEET]
            
            updated = False
            
            # Ищем все строки с данным user_id
            for row in range(2, ws.max_row + 1):  # начинаем с 2, пропуская заголовок
                if ws.cell(row, 3).value == user_id:  # колонка C - User ID
                    if username:
                        ws.cell(row, 4).value = username  # колонка D - Username
                    if fio:
                        ws.cell(row, 5).value = fio  # колонка E - ФИО
                    updated = True
            
            if updated:
                wb.save(self.filename)
                logger.info(f"Информация о пользователе {user_id} обновлена")
            else:
                logger.warning(f"Пользователь {user_id} не найден в Excel")
            
        except Exception as e:
            logger.error(f"Ошибка при обновлении информации пользователя: {e}")
    
    def get_all_messages(self) -> List[Dict]:
        """Получает все сообщения из Excel"""
        try:
            df = pd.read_excel(self.filename, sheet_name=EXCEL_SHEET)
            return df.to_dict('records')
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel: {e}")
            return []
    
    def get_user_messages(self, user_id: int) -> List[Dict]:
        """Получает все сообщения конкретного пользователя"""
        try:
            df = pd.read_excel(self.filename, sheet_name=EXCEL_SHEET)
            user_messages = df[df['User ID'] == user_id].to_dict('records')
            return user_messages
        except Exception as e:
            logger.error(f"Ошибка при чтении сообщений пользователя: {e}")
            return []
    
    def export_to_file(self) -> str:
        """Экспортирует данные в файл для отправки"""
        return self.filename
    
    def export_user_data(self, user_id: int):
        """Экспортирует данные конкретного пользователя в отдельный файл"""
        try:
            df = pd.read_excel(self.filename, sheet_name=EXCEL_SHEET)
            user_df = df[df['User ID'] == user_id]
            
            if user_df.empty:
                logger.warning(f"Нет данных для пользователя {user_id}")
                return None, None
            
            # Получаем ФИО пользователя
            user_fio = user_df['ФИО'].iloc[0] if not user_df['ФИО'].isna().all() else f"user_{user_id}"
            
            # Очищаем ФИО для имени файла
            safe_fio = "".join(c for c in user_fio if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_fio = safe_fio.replace(' ', '_')
            
            # Создаем имя файла с датой
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{safe_fio}_data_{timestamp}.xlsx"
            
            # Создаем временную папку для экспортов
            export_folder = "exports"
            Path(export_folder).mkdir(exist_ok=True)
            
            filepath = os.path.join(export_folder, filename)
            
            # Сохраняем файл
            user_df.to_excel(filepath, index=False, sheet_name=EXCEL_SHEET)
            logger.info(f"Экспортированы данные пользователя {user_fio} (ID: {user_id}) в {filepath}")
            
            return filepath, user_fio
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте данных пользователя: {e}")
            return None, None
    
    def get_all_users(self) -> List[Dict]:
        """Получает список всех пользователей с их данными"""
        try:
            df = pd.read_excel(self.filename, sheet_name=EXCEL_SHEET)
            
            users = []
            for user_id in df['User ID'].unique():
                user_df = df[df['User ID'] == user_id]
                fio = user_df['ФИО'].iloc[0] if not user_df['ФИО'].isna().all() else "Не указано"
                
                users.append({
                    'user_id': int(user_id),
                    'fio': fio,
                    'message_count': len(user_df),
                    'first_message': user_df['Дата и время'].min(),
                    'last_message': user_df['Дата и время'].max()
                })
            
            return sorted(users, key=lambda x: x['fio'])
            
        except Exception as e:
            logger.error(f"Ошибка при получении списка пользователей: {e}")
            return []
    
    def get_statistics(self) -> Dict:
        """Получает статистику по сообщениям"""
        try:
            df = pd.read_excel(self.filename, sheet_name=EXCEL_SHEET)
            
            stats = {
                'total_messages': len(df),
                'text_messages': len(df[df['Тип'] == 'text']),
                'photo_messages': len(df[df['Тип'] == 'photo']),
                'unique_users': df['User ID'].nunique(),
                'users': []
            }
            
            # Статистика по пользователям
            for user_id in df['User ID'].unique():
                user_df = df[df['User ID'] == user_id]
                stats['users'].append({
                    'user_id': user_id,
                    'fio': user_df['ФИО'].iloc[0] if not user_df['ФИО'].isna().all() else 'Не указано',
                    'messages_count': len(user_df),
                    'first_message': user_df['Дата и время'].min(),
                    'last_message': user_df['Дата и время'].max()
                })
            
            return stats
        except Exception as e:
            logger.error(f"Ошибка при получении статистики: {e}")
            return {}
    
    def create_manual_backup(self) -> Optional[str]:
        """Создает ручной бэкап данных"""
        return self._create_backup("manual")
    
    def list_backups(self) -> List[Dict]:
        """Получает список всех бэкапов"""
        try:
            backups = []
            if os.path.exists(BACKUP_FOLDER):
                for file in os.listdir(BACKUP_FOLDER):
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(BACKUP_FOLDER, file)
                        stat = os.stat(file_path)
                        backups.append({
                            'name': file,
                            'path': file_path,
                            'size': stat.st_size,
                            'created': datetime.fromtimestamp(stat.st_ctime),
                            'modified': datetime.fromtimestamp(stat.st_mtime)
                        })
            return sorted(backups, key=lambda x: x['created'], reverse=True)
        except Exception as e:
            logger.error(f"Ошибка при получении списка бэкапов: {e}")
            return []
    
    def restore_backup(self, backup_name: str) -> bool:
        """Восстанавливает данные из бэкапа"""
        try:
            backup_path = os.path.join(BACKUP_FOLDER, backup_name)
            if os.path.exists(backup_path):
                # Создаем бэкап текущего файла перед восстановлением
                self._create_backup("pre_restore")
                
                # Восстанавливаем
                import shutil
                shutil.copy2(backup_path, self.filename)
                logger.info(f"Восстановлен бэкап: {backup_name}")
                return True
            else:
                logger.error(f"Бэкап не найден: {backup_name}")
                return False
        except Exception as e:
            logger.error(f"Ошибка при восстановлении бэкапа: {e}")
            return False
    
    def cleanup_old_backups(self):
        """Удаляет старые бэкапы"""
        try:
            if not os.path.exists(BACKUP_FOLDER):
                return
            
            cutoff_time = datetime.now() - timedelta(days=BACKUP_RETENTION_DAYS)
            deleted_count = 0
            
            for file in os.listdir(BACKUP_FOLDER):
                if file.endswith('.xlsx'):
                    file_path = os.path.join(BACKUP_FOLDER, file)
                    file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                    
                    if file_time < cutoff_time:
                        os.remove(file_path)
                        deleted_count += 1
                        logger.info(f"Удален старый бэкап: {file}")
            
            if deleted_count > 0:
                logger.info(f"Удалено {deleted_count} старых бэкапов")
                
        except Exception as e:
            logger.error(f"Ошибка при очистке старых бэкапов: {e}")

# ========== YANDEX DISK ==========
class YandexDiskUploader:
    def __init__(self, token: str):
        self.y = yadisk.YaDisk(token=token)
        self.current_folder = BASE_YANDEX_FOLDER
        self.check_connection()
    
    def check_connection(self):
        """Проверка подключения к Яндекс.Диску"""
        try:
            if not self.y.check_token():
                raise Exception("Недействительный токен Яндекс.Диска")
            
            # Проверяем/создаем основную папку
            if not self.y.exists(BASE_YANDEX_FOLDER):
                self.y.mkdir(BASE_YANDEX_FOLDER)
                logger.info(f"Создана основная папка {BASE_YANDEX_FOLDER}")
            
            # Создаем папку для бэкапов
            backup_folder = f"{BASE_YANDEX_FOLDER}/_backups"
            if not self.y.exists(backup_folder):
                self.y.mkdir(backup_folder)
                logger.info(f"Создана папка для бэкапов {backup_folder}")
            
            # Создаем папку для экспортов
            exports_folder = f"{BASE_YANDEX_FOLDER}/_exports"
            if not self.y.exists(exports_folder):
                self.y.mkdir(exports_folder)
                logger.info(f"Создана папка для экспортов {exports_folder}")
            
            logger.info("Яндекс.Диск подключен успешно")
            return True
        except Exception as e:
            logger.error(f"Ошибка подключения к Яндекс.Диску: {e}")
            return False
    
    def set_current_folder(self, folder_path: str):
        """Устанавливает текущую папку для загрузки"""
        self.current_folder = folder_path
    
    def get_current_folder(self):
        """Получает текущую папку"""
        return self.current_folder
    
    def create_folder(self, folder_name: str, parent_folder: str = None) -> Optional[str]:
        """Создает новую папку и возвращает путь к созданной папке"""
        try:
            if parent_folder is None:
                parent_folder = self.current_folder
            
            new_folder_path = f"{parent_folder}/{folder_name}"
            
            if not self.y.exists(new_folder_path):
                self.y.mkdir(new_folder_path)
                logger.info(f"Создана папка: {new_folder_path}")
                return new_folder_path
            else:
                logger.warning(f"Папка уже существует: {new_folder_path}")
                return new_folder_path
        except Exception as e:
            logger.error(f"Ошибка создания папки: {e}")
            return None
    
    def list_folders(self, folder_path: str = None) -> List[Dict[str, Any]]:
        """Получает список папок в указанной директории"""
        try:
            if folder_path is None:
                folder_path = self.current_folder
            
            if not self.y.exists(folder_path):
                return []
            
            items = list(self.y.listdir(folder_path))
            folders = []
            
            for item in items:
                if item.type == 'dir' and not item.name.startswith('_'):
                    folder_data = {
                        'name': item.name,
                        'path': item.path,
                        'type': item.type
                    }
                    folders.append(folder_data)
            
            return folders
        except Exception as e:
            logger.error(f"Ошибка получения списка папок: {e}")
            return []
    
    def get_parent_folder(self, folder_path: str) -> Optional[str]:
        """Получает родительскую папку"""
        if folder_path == BASE_YANDEX_FOLDER:
            return None
        return os.path.dirname(folder_path.rstrip('/'))
    
    def upload_file(self, file_path: str, file_name: str) -> Optional[str]:
        """Загружает файл в текущую папку Яндекс.Диска"""
        try:
            remote_path = f"{self.current_folder}/{file_name}"
            
            # Если файл уже существует, добавляем timestamp
            if self.y.exists(remote_path):
                name, ext = os.path.splitext(file_name)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"{name}_{timestamp}{ext}"
                remote_path = f"{self.current_folder}/{file_name}"
            
            # Загружаем файл
            self.y.upload(file_path, remote_path)
            
            # Получаем публичную ссылку
            self.y.publish(remote_path)
            resource = self.y.get_meta(remote_path)
            public_url = resource.public_url
            
            logger.info(f"Файл загружен: {file_name} в папку {self.current_folder}")
            logger.info(f"Публичная ссылка: {public_url}")
            
            return public_url
        
        except yadisk.exceptions.PathExistsError:
            logger.error(f"Файл уже существует: {file_name}")
            return None
        except Exception as e:
            logger.error(f"Ошибка загрузки в Яндекс.Диск: {e}")
            return None
    
    def upload_backup(self, local_path: str) -> Optional[str]:
        """Загружает бэкап на Яндекс.Диск"""
        try:
            backup_name = os.path.basename(local_path)
            remote_path = f"{BASE_YANDEX_FOLDER}/_backups/{backup_name}"
            
            if not self.y.exists(f"{BASE_YANDEX_FOLDER}/_backups"):
                self.y.mkdir(f"{BASE_YANDEX_FOLDER}/_backups")
            
            self.y.upload(local_path, remote_path)
            logger.info(f"Бэкап загружен на Яндекс.Диск: {remote_path}")
            return remote_path
        except Exception as e:
            logger.error(f"Ошибка загрузки бэкапа на Яндекс.Диск: {e}")
            return None
    
    def upload_export_to_user_folder(self, local_path: str, user_fio: str):
        """Загружает экспорт данных пользователя в его существующую папку на Яндекс.Диске"""
        try:
            # Очищаем ФИО для имени папки
            safe_fio = "".join(c for c in user_fio if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_fio = safe_fio.replace(' ', '_')
            
            # Путь к папке пользователя (не создаем новую, если её нет)
            user_folder = f"{BASE_YANDEX_FOLDER}/{safe_fio}"
            
            # Проверяем, существует ли папка пользователя
            if not self.y.exists(user_folder):
                logger.warning(f"Папка пользователя {user_folder} не существует, экспорт будет загружен в папку _exports")
                user_folder = f"{BASE_YANDEX_FOLDER}/_exports"
                if not self.y.exists(user_folder):
                    self.y.mkdir(user_folder)
            
            file_name = os.path.basename(local_path)
            remote_path = f"{user_folder}/{file_name}"
            
            # Если файл с таким именем уже существует, добавляем timestamp
            if self.y.exists(remote_path):
                name, ext = os.path.splitext(file_name)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"{name}_{timestamp}{ext}"
                remote_path = f"{user_folder}/{file_name}"
            
            self.y.upload(local_path, remote_path)
            logger.info(f"Экспорт загружен в папку пользователя: {remote_path}")
            
            # Получаем публичную ссылку
            self.y.publish(remote_path)
            resource = self.y.get_meta(remote_path)
            public_url = resource.public_url
            
            return public_url, user_folder
            
        except Exception as e:
            logger.error(f"Ошибка загрузки экспорта в папку пользователя: {e}")
            return None, None
    
    def get_disk_info(self):
        """Получение информации о диске"""
        try:
            disk_info = self.y.get_disk_info()
            return {
                'total': disk_info.total_space,
                'used': disk_info.used_space,
                'free': disk_info.free_space
            }
        except Exception as e:
            logger.error(f"Ошибка получения информации о диске: {e}")
            return None
    
    def get_public_url(self, folder_path: str) -> Optional[str]:
        """Получает публичную ссылку на папку"""
        try:
            # Публикуем папку если она еще не опубликована
            self.y.publish(folder_path)
            resource = self.y.get_meta(folder_path)
            return resource.public_url
        except Exception as e:
            logger.error(f"Ошибка получения публичной ссылки на папку: {e}")
            return None

# ========== TELEGRAM BOT ==========
# Инициализация бота с правильными параметрами
bot = Bot(
    token=TELEGRAM_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
router = Router()
dp.include_router(router)

# Инициализируем Яндекс.Диск
try:
    disk_uploader = YandexDiskUploader(YANDEX_TOKEN)
except Exception as e:
    logger.error(f"Не удалось инициализировать Яндекс.Диск: {e}")
    disk_uploader = None

# Инициализируем Excel менеджер
excel_manager = ExcelManager()

# Папка для временного хранения файлов
TEMP_FOLDER = "temp_photos"
Path(TEMP_FOLDER).mkdir(exist_ok=True)

# Папка для хранения файлов гайда
GUIDE_FOLDER = "guides"
Path(GUIDE_FOLDER).mkdir(exist_ok=True)

# Папка для экспортов
EXPORT_FOLDER = "exports"
Path(EXPORT_FOLDER).mkdir(exist_ok=True)

# Хранилище для медиагрупп
media_group_storage = defaultdict(list)

# Хранилище для путей папок по индексам
folder_paths = {}
folder_index_counter = 0

# Флаг для планировщика бэкапов
backup_task_running = False

# ========== КЛАВИАТУРЫ ==========
def get_main_keyboard() -> ReplyKeyboardMarkup:
    """Создает основную клавиатуру с кнопками"""
    keyboard = [
        [
            KeyboardButton(text="🏠 Стартуем"),
            KeyboardButton(text="📂 Выбрать папку для загрузки")
        ],
        [
            KeyboardButton(text="📁 Открыть основную папку"),
            KeyboardButton(text="📖 Гайд по боту")
        ],
        [
            KeyboardButton(text="📊 Показать данные"),
            KeyboardButton(text="📊 Статистика"),
            KeyboardButton(text="👤 Мои данные")
        ],
        [
            KeyboardButton(text="💾 Управление бэкапами"),
            KeyboardButton(text="📤 Экспорт пользователя")
        ],
        [
            KeyboardButton(text="❓ Помощь")
        ],
        [
            KeyboardButton(text="❌ Отмена")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите действие..."
    )

def get_after_photos_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру после загрузки фото (с возможностью пропуска текста)"""
    keyboard = [
        [
            KeyboardButton(text="📸 Загрузить ещё фото"),
            KeyboardButton(text="📝 Отправить текст")
        ],
        [
            KeyboardButton(text="⏭️ Пропустить текст (завершить)")
        ],
        [
            KeyboardButton(text="🏠 В главное меню")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите действие..."
    )

def get_skip_text_confirmation_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру для подтверждения пропуска текста"""
    keyboard = [
        [
            KeyboardButton(text="✅ Да, пропустить"),
            KeyboardButton(text="❌ Нет, ввести текст")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите действие..."
    )

def get_after_folder_selection_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру после выбора папки (до загрузки фото)"""
    keyboard = [
        [
            KeyboardButton(text="📸 Загрузить фото")
        ],
        [
            KeyboardButton(text="📁 Текущая папка"),
            KeyboardButton(text="📂 Основная папка (Фото за МК)")
        ],
        [
            KeyboardButton(text="📂 Выбрать другую папку"),
            KeyboardButton(text="🏠 В главное меню")
        ],
        [
            KeyboardButton(text="❌ Отмена")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите действие..."
    )

def get_after_text_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру после отправки текста (возврат к начальным кнопкам)"""
    return get_main_keyboard()

def get_text_options_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру для опций текста (без кнопок)"""
    return ReplyKeyboardRemove()

def get_upload_options_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру с опциями загрузки"""
    keyboard = [
        [
            KeyboardButton(text="📝 Дать своё название"),
            KeyboardButton(text="🔄 Использовать автоназвание")
        ],
        [
            KeyboardButton(text="📁 Загрузить все с автоназваниями"),
            KeyboardButton(text="❌ Отмена")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите опцию..."
    )

def get_album_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру для режима альбома"""
    keyboard = [
        [KeyboardButton(text="✅ Готово")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Отправляйте фото по одному..."
    )

def get_batch_upload_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру для пакетной загрузки"""
    keyboard = [
        [
            KeyboardButton(text="📝 Дать название каждому"),
            KeyboardButton(text="🔄 Все с автоназваниями")
        ],
        [
            KeyboardButton(text="❌ Отмена")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите способ загрузки..."
    )

def get_cancel_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру только с кнопкой отмены"""
    keyboard = [[KeyboardButton(text="❌ Отмена")]]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Используйте /cancel или кнопку для отмены"
    )

def get_export_user_keyboard(users: List[Dict]) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора пользователя для экспорта"""
    keyboard = []
    
    for user in users[:20]:  # Показываем не более 20 пользователей
        display_name = user['fio'][:30] + "..." if len(user['fio']) > 30 else user['fio']
        keyboard.append([
            InlineKeyboardButton(
                text=f"👤 {display_name} ({user['message_count']} сообщ.)",
                callback_data=f"export_user:{user['user_id']}"
            )
        ])
    
    if len(users) > 20:
        keyboard.append([
            InlineKeyboardButton(
                text=f"📋 Всего пользователей: {len(users)} (показаны первые 20)",
                callback_data="noop"
            )
        ])
    
    keyboard.append([
        InlineKeyboardButton(text="🏠 В главное меню", callback_data="back_to_main_menu")
    ])
    
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

def get_backup_keyboard() -> InlineKeyboardMarkup:
    """Создает клавиатуру для управления бэкапами"""
    keyboard = [
        [InlineKeyboardButton(text="💾 Создать бэкап", callback_data="create_backup")],
        [InlineKeyboardButton(text="📋 Список бэкапов", callback_data="list_backups")],
        [InlineKeyboardButton(text="🗑️ Очистить старые бэкапы", callback_data="cleanup_backups")],
        [InlineKeyboardButton(text="🏠 В главное меню", callback_data="back_to_main_menu")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

def format_folder_name(name: str, max_length: int = 15) -> str:
    """Форматирует название папки для отображения в кнопке"""
    if len(name) <= max_length:
        return name
    return name[:max_length-3] + "..."

def get_folder_keyboard(folders: list, current_path: str, is_root: bool = False) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора папки с навигацией (2 колонки)"""
    global folder_index_counter
    keyboard = []
    
    # Кнопка "Наверх" если не в корне (занимает всю ширину)
    if not is_root:
        parent_path = disk_uploader.get_parent_folder(current_path)
        if parent_path:
            # Сохраняем путь родительской папки
            parent_index = folder_index_counter
            folder_paths[parent_index] = parent_path
            folder_index_counter += 1
            
            keyboard.append([InlineKeyboardButton(
                text="⬆️ Наверх", 
                callback_data=f"nav_folder:{parent_index}"
            )])
    
    # Добавляем папки в два столбца
    folder_buttons = []
    for i, folder in enumerate(folders):
        folder_name = folder.get('name', 'Без названия')
        folder_path = folder.get('path', '')
        
        # Сохраняем путь папки
        folder_index = folder_index_counter
        folder_paths[folder_index] = folder_path
        folder_index_counter += 1
        
        # Форматируем название для кнопки
        display_name = format_folder_name(folder_name, 15)
        
        # Создаем кнопку
        button = InlineKeyboardButton(
            text=f"📁 {display_name}", 
            callback_data=f"nav_folder:{folder_index}"
        )
        
        folder_buttons.append(button)
        
        # Каждые 2 кнопки добавляем ряд
        if len(folder_buttons) == 2:
            keyboard.append(folder_buttons)
            folder_buttons = []
    
    # Добавляем оставшиеся кнопки (если есть)
    if folder_buttons:
        keyboard.append(folder_buttons)
    
    # Кнопка выбора текущей папки (в два столбца с кнопкой создания)
    current_index = folder_index_counter
    folder_paths[current_index] = current_path
    folder_index_counter += 1
    
    keyboard.append([
        InlineKeyboardButton(text="✅ Выбрать эту папку", callback_data=f"select_folder:{current_index}"),
        InlineKeyboardButton(text="➕ Создать папку", callback_data="create_new_folder")
    ])
    
    # Кнопки действий (в два столбца)
    keyboard.append([
        InlineKeyboardButton(text="🏠 В главное меню", callback_data="back_to_main_menu"),
        InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_folder_selection")
    ])
    
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

def get_folder_creation_keyboard(folder_path: str, folder_index: int) -> InlineKeyboardMarkup:
    """Создает клавиатуру после создания папки"""
    # Сохраняем путь во временном хранилище и передаем только индекс
    folder_paths[folder_index] = folder_path
    
    keyboard = [
        [InlineKeyboardButton(text="✅ Перейти в созданную папку", callback_data=f"select_new_folder:{folder_index}")],
        [InlineKeyboardButton(text="📂 Остаться в текущей", callback_data="select_current")],
        [InlineKeyboardButton(text="🔙 Назад к меню", callback_data="back_to_main_menu")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

def get_folder_link_keyboard(folder_path: str, public_url: str = None, base_public_url: str = None) -> InlineKeyboardMarkup:
    """Создает клавиатуру со ссылками на папки"""
    keyboard = []
    
    if public_url:
        keyboard.append([InlineKeyboardButton(text="📁 Текущая папка", url=public_url)])
    
    if base_public_url:
        keyboard.append([InlineKeyboardButton(text="📂 Основная папка (Фото за МК)", url=base_public_url)])
    
    keyboard.append([InlineKeyboardButton(text="📂 Выбрать другую папку", callback_data="select_another_folder")])
    keyboard.append([InlineKeyboardButton(text="🏠 В главное меню", callback_data="back_to_main_menu")])
    
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

def get_main_folder_keyboard() -> InlineKeyboardMarkup:
    """Создает клавиатуру со ссылкой на основную папку"""
    keyboard = []
    
    keyboard.append([InlineKeyboardButton(text="📂 Открыть основную папку (Фото за МК)", url=BASE_FOLDER_URL)])
    keyboard.append([InlineKeyboardButton(text="🏠 Вернуться в меню", callback_data="back_to_main_menu")])
    
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

def get_photo_upload_keyboard() -> ReplyKeyboardMarkup:
    """Создает клавиатуру для загрузки фото после выбора папки"""
    keyboard = [
        [
            KeyboardButton(text="📸 Начать загрузку фото"),
            KeyboardButton(text="❌ Отмена")
        ]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        input_field_placeholder="Выберите действие..."
    )

def get_guide_keyboard() -> InlineKeyboardMarkup:
    """Создает клавиатуру с ссылками на гайд"""
    keyboard = [
        [InlineKeyboardButton(text="📖 Открыть гайд онлайн", url=GUIDE_URL)],
        [InlineKeyboardButton(text="📥 Скачать гайд", callback_data="download_guide")],
        [InlineKeyboardButton(text="🏠 В главное меню", callback_data="back_to_main_menu")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

# ========== BACKUP SCHEDULER ==========
async def backup_scheduler():
    """Планировщик автоматического резервного копирования"""
    global backup_task_running
    
    if backup_task_running:
        return
    
    backup_task_running = True
    logger.info("Планировщик бэкапов запущен")
    
    while True:
        try:
            # Ждем указанный интервал
            await asyncio.sleep(BACKUP_INTERVAL_HOURS * 60 * 60)
            
            logger.info("Запуск автоматического создания бэкапа...")
            
            # Создаем бэкап
            backup_path = excel_manager.create_manual_backup()
            
            if backup_path:
                logger.info(f"Автоматический бэкап создан: {backup_path}")
                
                # Загружаем бэкап на Яндекс.Диск если доступен
                if disk_uploader:
                    disk_uploader.upload_backup(backup_path)
                    logger.info("Бэкап загружен на Яндекс.Диск")
                
                # Очищаем старые бэкапы
                excel_manager.cleanup_old_backups()
                logger.info("Очистка старых бэкапов выполнена")
            else:
                logger.warning("Не удалось создать автоматический бэкап")
                
        except Exception as e:
            logger.error(f"Ошибка в планировщике бэкапов: {e}")
            # В случае ошибки ждем 1 час и пробуем снова
            await asyncio.sleep(60 * 60)

# ========== HANDLERS ==========
# Глобальные переменные для хранения данных
current_folder_list = []
user_photo_data = defaultdict(list)  # Для хранения списка фото перед загрузкой
user_album_data = defaultdict(dict)  # Для хранения данных об альбоме
user_batch_data = defaultdict(dict)  # Для хранения данных о пакетной загрузке
user_text_data = defaultdict(dict)  # Для хранения текстовых данных
user_fio_data = defaultdict(str)  # Для хранения ФИО пользователя
user_has_uploaded_photos = defaultdict(bool)  # Флаг, загружал ли пользователь фото
user_has_sent_text = defaultdict(bool)  # Флаг, отправлял ли пользователь текст
user_skip_text_requested = defaultdict(bool)  # Флаг, запросил ли пользователь пропуск текста

@router.message(Command("start", "help"))
@router.message(F.text == "🏠 Стартуем")
async def send_welcome(message: Message, state: FSMContext):
    """Обработчик команды /start и кнопки Старт"""
    await state.clear()
    
    current_folder = disk_uploader.get_current_folder() if disk_uploader else "Не доступно"
    
    welcome_text = (
        "🏠 <b>Бот для загрузки фото с МК в Яндекс.Диск</b>\n\n"
        "Я вас категорически приветствую, молодые люди (или уже не молодые).\n\n"
        "Этот бот был создан, чтобы загружать в него фотографии с мастер-классов, которые сразу будут загружены в Яндекс.Диск (облегчил жизнь для вас, так скажем).\n"
        "📸 <b>Порядок работы с фотографиями:</b>\n"
        "1. Сначала выбери папку для загрузки (можно создать новую или выбрать существующую)\n"
        "2. После выбора папки можно загрузить фотографии\n\n"
        "📝 <b>Текстовые сообщения (опционально):</b>\n"
        "После загрузки фотографий ты можешь отправить текст с описанием МК.\n"
        "Для этого нажми '📝 Отправить текст'.\n"
        "Если не хочешь отправлять текст, нажми '⏭️ Пропустить текст (завершить)'.\n\n"
        "💾 <b>Резервное копирование:</b>\n"
        f"• Автоматические бэкапы создаются каждые {BACKUP_INTERVAL_HOURS} часов\n"
        f"• Бэкапы хранятся {BACKUP_RETENTION_DAYS} дней\n"
        "• Используйте кнопку '💾 Управление бэкапами' для ручного управления\n\n"
        "👤 <b>Экспорт данных:</b>\n"
        "• '👤 Мои данные' - выгрузить ваши личные данные\n"
        "• '📤 Экспорт пользователя' - выгрузить данные любого пользователя (доступно администратору)\n\n"
        "Помимо этого вы можете создать свою папку с помощью бота и туда выгружать фотографии (если не будет папки с вашей фамилией и именем).\n\n"
        "И помните: я слежу за вами.\n\n"
        "Счастливых голодных игр\n\n"
        f"📁 Основная папка: <code>{BASE_YANDEX_FOLDER}</code>\n"
        f"📂 Текущая папка: <code>{current_folder}</code>\n\n"
        f"🔗 <a href='{BASE_FOLDER_URL}'>Открыть основную папку в Яндекс.Диске</a>\n\n"
        "<b>Используйте кнопки ниже или команды:</b>\n"
        "/folder - выбрать папку для загрузки\n"
        "/newfolder - создать новую папку\n"
        "/open_main - открыть основную папку\n"
        "/guide - открыть гайд по боту\n"
        "/export - выгрузить все данные в Excel\n"
        "/export_me - выгрузить мои данные\n"
        "/stats - показать статистику\n"
        "/backup - управление бэкапами\n"
        "/cancel - отменить текущую операцию (до загрузки фото)"
    )
    await message.answer(welcome_text, reply_markup=get_main_keyboard(), parse_mode=ParseMode.HTML)

@router.message(Command("guide"))
@router.message(F.text == "📖 Гайд по боту")
async def send_guide(message: Message, state: FSMContext):
    """Обработчик отправки гайда по боту"""
    try:
        # Получаем путь к папке, где находится скрипт
        script_dir = Path(__file__).parent
        
        # Ищем файл гайда в различных местах
        possible_paths = [
            script_dir / "Гайд на тг бота.pdf",
            script_dir / "guide.pdf",
            script_dir / "Гайд.pdf",
            script_dir / GUIDE_FOLDER / "Гайд на тг бота.pdf",
            script_dir / GUIDE_FOLDER / "guide.pdf",
            Path("Гайд на тг бота.pdf"),
            Path("guide.pdf")
        ]
        
        guide_path = None
        for path in possible_paths:
            if path.exists():
                guide_path = path
                break
        
        if guide_path and guide_path.exists():
            # Отправляем PDF-файл
            document = FSInputFile(guide_path)
            await message.answer_document(
                document,
                caption="📖 <b>Гайд по использованию бота</b>\n\n"
                        "Здесь вы найдете подробную инструкцию по работе с ботом.\n"
                        "Скачайте и читайте в любое удобное время!",
                parse_mode=ParseMode.HTML
            )
            logger.info(f"Гайд отправлен пользователю {message.from_user.id}")
        else:
            # Если файл не найден, отправляем текстовую версию гайда с ссылкой
            guide_text = (
                "📖 <b>Гайд по использованию бота</b>\n\n"
                "<b>📸 Загрузка фотографий:</b>\n"
                "1. Нажми кнопку '📂 Выбрать папку для загрузки'\n"
                "2. Выбери существующую папку или создайте новую (кнопка '➕ Создать папку')\n"
                "3. Нажми '✅ Выбрать эту папку'\n"
                "4. Нажми '📸 Загрузить фото'\n"
                "5. Отправь фотографии (можно несколько сразу)\n"
                "6. Выбери способ именования файлов:\n"
                "   • '📝 Дать название каждому' - ввести название для каждого фото\n"
                "   • '🔄 Все с автоназваниями' - автоматические названия\n\n"
                "<b>📝 Отправка текста (опционально):</b>\n"
                "1. После загрузки фото нажми '📝 Отправить текст'\n"
                "2. Введи ваше ФИО\n"
                "3. Введи название МК\n"
                "4. Введи его описание\n"
                "5. Или нажми '⏭️ Пропустить текст (завершить)' чтобы завершить без текста\n\n"
                "<b>📊 Просмотр данных:</b>\n"
                "• '📊 Показать данные' - выгрузить Excel файл со всеми записями\n"
                "• '📊 Статистика' - показать статистику по всем сообщениям\n"
                "• '👤 Мои данные' - выгрузить только ваши данные\n"
                "• '📤 Экспорт пользователя' - выгрузить данные другого пользователя\n\n"
                "<b>💾 Управление бэкапами:</b>\n"
                "• '💾 Управление бэкапами' - создать бэкап, просмотреть список, очистить старые\n"
                f"• Автоматические бэкапы создаются каждые {BACKUP_INTERVAL_HOURS} часов\n"
                f"• Бэкапы хранятся {BACKUP_RETENTION_DAYS} дней\n\n"
                "<b>📁 Работа с папками:</b>\n"
                "• '📁 Текущая папка' - открыть текущую папку\n"
                "• '📂 Основная папка' - открыть основную папку\n"
                "• '📂 Выбрать другую папку' - изменить папку для загрузки\n\n"
                "<b>❓ Дополнительно:</b>\n"
                "• '🏠 Стартуем' - вернуться в главное меню\n"
                "• '❌ Отмена' - отменить текущую операцию\n\n"
                "📥 <b>Скачать полную версию гайда в PDF:</b>\n"
                "Нажмите кнопку ниже '📥 Скачать гайд'"
            )
            
            await message.answer(
                guide_text,
                parse_mode=ParseMode.HTML,
                reply_markup=get_guide_keyboard()
            )
            
    except Exception as e:
        logger.error(f"Ошибка при отправке гайда: {e}")
        await message.answer(
            "❌ Произошла ошибка при отправке гайда.\n\n"
            "Пожалуйста, попробуйте позже или обратитесь к администратору.",
            reply_markup=get_main_keyboard()
        )

@router.message(Command("export_me"))
@router.message(F.text == "👤 Мои данные")
async def export_my_data(message: Message, state: FSMContext):
    """Экспорт данных текущего пользователя"""
    user_id = message.from_user.id
    
    status_msg = await message.answer("⏳ Подготавливаю ваши данные...")
    
    try:
        # Получаем данные пользователя
        user_messages = excel_manager.get_user_messages(user_id)
        
        if not user_messages:
            await status_msg.edit_text(
                "📊 <b>Нет данных для экспорта</b>\n\n"
                "У вас пока нет сохраненных сообщений.\n"
                "Сначала загрузите фото и отправьте текст.",
                parse_mode=ParseMode.HTML
            )
            return
        
        # Получаем ФИО пользователя
        fio = user_fio_data.get(user_id, "Не указано")
        for msg in user_messages:
            if msg.get('ФИО') and msg['ФИО'] != "Не указано":
                fio = msg['ФИО']
                break
        
        # Создаем файл экспорта
        export_path, user_fio = excel_manager.export_user_data(user_id)
        
        if export_path and os.path.exists(export_path):
            # Отправляем файл пользователю
            document = FSInputFile(export_path)
            
            await status_msg.delete()
            
            # Загружаем в папку пользователя на Яндекс.Диске
            folder_path = None
            if disk_uploader:
                public_url, folder_path = disk_uploader.upload_export_to_user_folder(export_path, fio)
                if public_url:
                    await message.answer_document(
                        document,
                        caption=f"👤 <b>Ваши данные</b>\n\n"
                                f"ФИО: <code>{fio}</code>\n"
                                f"Всего записей: {len(user_messages)}\n"
                                f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                                f"📁 Файл сохранен в папку: <code>{folder_path}</code>\n"
                                f"🔗 <a href='{public_url}'>Открыть файл на Яндекс.Диске</a>",
                        parse_mode=ParseMode.HTML,
                        reply_markup=get_main_keyboard()
                    )
                else:
                    await message.answer_document(
                        document,
                        caption=f"👤 <b>Ваши данные</b>\n\n"
                                f"ФИО: <code>{fio}</code>\n"
                                f"Всего записей: {len(user_messages)}\n"
                                f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                                f"❌ Не удалось загрузить файл на Яндекс.Диск.",
                        parse_mode=ParseMode.HTML,
                        reply_markup=get_main_keyboard()
                    )
            else:
                await message.answer_document(
                    document,
                    caption=f"👤 <b>Ваши данные</b>\n\n"
                            f"ФИО: <code>{fio}</code>\n"
                            f"Всего записей: {len(user_messages)}\n"
                            f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    parse_mode=ParseMode.HTML,
                    reply_markup=get_main_keyboard()
                )
            
            # Удаляем временный файл
            try:
                os.remove(export_path)
            except:
                pass
            
            logger.info(f"Экспортированы данные пользователя {user_id} (ФИО: {fio})")
        else:
            await status_msg.edit_text(
                "❌ Ошибка при экспорте данных.",
                parse_mode=ParseMode.HTML
            )
            
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных пользователя: {e}")
        await status_msg.edit_text(
            "❌ Ошибка при экспорте данных.",
            parse_mode=ParseMode.HTML
        )

@router.message(F.text == "📤 Экспорт пользователя")
async def export_user_data(message: Message, state: FSMContext):
    """Экспорт данных выбранного пользователя (административная функция)"""
    # Проверяем, является ли пользователь администратором
    ADMIN_IDS = [2138391498]  # Замените на реальные ID администраторов
    
    user_id = message.from_user.id
    
    if user_id not in ADMIN_IDS:
        await message.answer(
            "🔒 <b>Доступ ограничен</b>\n\n"
            "Эта функция доступна только администраторам бота.\n"
            "Для экспорта своих данных используйте кнопку '👤 Мои данные'.",
            parse_mode=ParseMode.HTML,
            reply_markup=get_main_keyboard()
        )
        return
    
    # Получаем список всех пользователей
    users = excel_manager.get_all_users()
    
    if not users:
        await message.answer(
            "📊 <b>Нет пользователей для экспорта</b>\n\n"
            "Пока нет сохраненных данных.",
            parse_mode=ParseMode.HTML,
            reply_markup=get_main_keyboard()
        )
        return
    
    # Показываем клавиатуру с выбором пользователя
    await message.answer(
        "👥 <b>Выберите пользователя для экспорта данных</b>\n\n"
        f"Всего пользователей: {len(users)}\n\n"
        "Выберите из списка ниже:",
        reply_markup=get_export_user_keyboard(users),
        parse_mode=ParseMode.HTML
    )
    
    await state.set_state(ExportStates.waiting_for_user_selection)

@router.callback_query(lambda c: c.data.startswith('export_user:'))
async def export_selected_user_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик экспорта данных выбранного пользователя"""
    await callback_query.answer()
    
    try:
        target_user_id = int(callback_query.data.split(':')[1])
        
        await callback_query.message.edit_text("⏳ Подготавливаю данные пользователя...")
        
        # Получаем данные пользователя
        user_messages = excel_manager.get_user_messages(target_user_id)
        
        if not user_messages:
            await callback_query.message.edit_text(
                "❌ Нет данных для этого пользователя."
            )
            await state.clear()
            return
        
        # Получаем ФИО пользователя
        fio = "Не указано"
        for msg in user_messages:
            if msg.get('ФИО') and msg['ФИО'] != "Не указано":
                fio = msg['ФИО']
                break
        
        # Создаем файл экспорта
        export_path, user_fio = excel_manager.export_user_data(target_user_id)
        
        if export_path and os.path.exists(export_path):
            # Отправляем файл администратору
            document = FSInputFile(export_path)
            
            await callback_query.message.delete()
            
            # Загружаем в папку пользователя на Яндекс.Диске
            folder_path = None
            if disk_uploader:
                public_url, folder_path = disk_uploader.upload_export_to_user_folder(export_path, fio)
                if public_url:
                    await callback_query.message.answer_document(
                        document,
                        caption=f"👤 <b>Данные пользователя</b>\n\n"
                                f"ФИО: <code>{fio}</code>\n"
                                f"User ID: <code>{target_user_id}</code>\n"
                                f"Всего записей: {len(user_messages)}\n"
                                f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                                f"📁 Файл сохранен в папку: <code>{folder_path}</code>\n"
                                f"🔗 <a href='{public_url}'>Открыть файл на Яндекс.Диске</a>",
                        parse_mode=ParseMode.HTML,
                        reply_markup=get_main_keyboard()
                    )
                else:
                    await callback_query.message.answer_document(
                        document,
                        caption=f"👤 <b>Данные пользователя</b>\n\n"
                                f"ФИО: <code>{fio}</code>\n"
                                f"User ID: <code>{target_user_id}</code>\n"
                                f"Всего записей: {len(user_messages)}\n"
                                f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                                f"❌ Не удалось загрузить файл на Яндекс.Диск.\n"
                                f"Папка пользователя не найдена.",
                        parse_mode=ParseMode.HTML,
                        reply_markup=get_main_keyboard()
                    )
            else:
                await callback_query.message.answer_document(
                    document,
                    caption=f"👤 <b>Данные пользователя</b>\n\n"
                            f"ФИО: <code>{fio}</code>\n"
                            f"User ID: <code>{target_user_id}</code>\n"
                            f"Всего записей: {len(user_messages)}\n"
                            f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    parse_mode=ParseMode.HTML,
                    reply_markup=get_main_keyboard()
                )
            
            # Удаляем временный файл
            try:
                os.remove(export_path)
            except:
                pass
            
            logger.info(f"Экспортированы данные пользователя {target_user_id} (ФИО: {fio})")
        else:
            await callback_query.message.edit_text(
                "❌ Ошибка при экспорте данных."
            )
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных пользователя: {e}")
        await callback_query.message.edit_text(
            "❌ Ошибка при экспорте данных."
        )
        await state.clear()

@router.callback_query(lambda c: c.data == "download_guide")
async def download_guide_callback(callback_query: CallbackQuery):
    """Обработчик скачивания гайда"""
    try:
        await callback_query.answer()
        await callback_query.message.answer(
            "📥 <b>Скачивание гайда</b>\n\n"
            f"Ссылка для скачивания: {GUIDE_URL}\n\n"
            "Если ссылка не открывается, скопируйте её и вставьте в браузер.",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        logger.error(f"Ошибка при скачивании гайда: {e}")
        await callback_query.answer("❌ Ошибка при скачивании")

@router.message(Command("backup"))
@router.message(F.text == "💾 Управление бэкапами")
async def manage_backups(message: Message, state: FSMContext):
    """Обработчик управления бэкапами"""
    await message.answer(
        "💾 <b>Управление резервными копиями</b>\n\n"
        f"📅 Автоматические бэкапы: каждые {BACKUP_INTERVAL_HOURS} часов\n"
        f"💾 Хранение бэкапов: {BACKUP_RETENTION_DAYS} дней\n\n"
        "Выберите действие:",
        reply_markup=get_backup_keyboard(),
        parse_mode=ParseMode.HTML
    )

@router.callback_query(lambda c: c.data == "create_backup")
async def create_backup_callback(callback_query: CallbackQuery):
    """Обработчик создания бэкапа"""
    await callback_query.answer()
    
    status_msg = await callback_query.message.answer("⏳ Создаю бэкап...")
    
    try:
        backup_path = excel_manager.create_manual_backup()
        
        if backup_path:
            # Получаем размер файла
            size = os.path.getsize(backup_path)
            size_mb = size / (1024 * 1024)
            
            # Загружаем на Яндекс.Диск если доступен
            if disk_uploader:
                disk_uploader.upload_backup(backup_path)
            
            await status_msg.edit_text(
                f"✅ <b>Бэкап успешно создан!</b>\n\n"
                f"📁 Файл: {os.path.basename(backup_path)}\n"
                f"📦 Размер: {size_mb:.2f} МБ\n"
                f"📅 Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                f"Бэкап сохранен локально и на Яндекс.Диске.",
                parse_mode=ParseMode.HTML
            )
        else:
            await status_msg.edit_text("❌ Не удалось создать бэкап.")
            
    except Exception as e:
        logger.error(f"Ошибка при создании бэкапа: {e}")
        await status_msg.edit_text("❌ Ошибка при создании бэкапа.")

@router.callback_query(lambda c: c.data == "list_backups")
async def list_backups_callback(callback_query: CallbackQuery):
    """Обработчик просмотра списка бэкапов"""
    await callback_query.answer()
    
    try:
        backups = excel_manager.list_backups()
        
        if not backups:
            await callback_query.message.answer(
                "📋 <b>Список бэкапов</b>\n\n"
                "Нет сохраненных бэкапов.",
                parse_mode=ParseMode.HTML,
                reply_markup=get_backup_keyboard()
            )
            return
        
        # Формируем список бэкапов
        backup_list = "📋 <b>Список бэкапов</b>\n\n"
        for i, backup in enumerate(backups[:10], 1):  # Показываем последние 10
            size_mb = backup['size'] / (1024 * 1024)
            backup_list += (
                f"{i}. {backup['name']}\n"
                f"   📅 {backup['created'].strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"   📦 {size_mb:.2f} МБ\n\n"
            )
        
        if len(backups) > 10:
            backup_list += f"<i>... и еще {len(backups) - 10} бэкапов</i>"
        
        await callback_query.message.answer(
            backup_list,
            parse_mode=ParseMode.HTML,
            reply_markup=get_backup_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка при получении списка бэкапов: {e}")
        await callback_query.message.answer("❌ Ошибка при получении списка бэкапов.")

@router.callback_query(lambda c: c.data == "cleanup_backups")
async def cleanup_backups_callback(callback_query: CallbackQuery):
    """Обработчик очистки старых бэкапов"""
    await callback_query.answer()
    
    status_msg = await callback_query.message.answer("⏳ Очищаю старые бэкапы...")
    
    try:
        excel_manager.cleanup_old_backups()
        
        await status_msg.edit_text(
            f"✅ <b>Очистка старых бэкапов выполнена!</b>\n\n"
            f"Удалены бэкапы старше {BACKUP_RETENTION_DAYS} дней.\n\n"
            f"Используйте '📋 Список бэкапов' для просмотра оставшихся.",
            parse_mode=ParseMode.HTML,
            reply_markup=get_backup_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка при очистке бэкапов: {e}")
        await status_msg.edit_text("❌ Ошибка при очистке бэкапов.")

@router.message(Command("open_main"))
@router.message(F.text == "📁 Открыть основную папку")
async def open_main_folder(message: Message, state: FSMContext):
    """Обработчик открытия основной папки"""
    await message.answer(
        f"📂 <b>Основная папка</b>\n\n"
        f"Папка: <code>{BASE_YANDEX_FOLDER}</code>\n\n"
        f"Нажмите на кнопку ниже, чтобы открыть папку в Яндекс.Диске:",
        reply_markup=get_main_folder_keyboard(),
        parse_mode=ParseMode.HTML
    )

@router.message(F.text == "📁 Текущая папка")
async def open_current_folder(message: Message, state: FSMContext):
    """Обработчик открытия текущей папки"""
    current_folder = disk_uploader.get_current_folder()
    current_public_url = disk_uploader.get_public_url(current_folder)
    
    # Определяем, какую клавиатуру показывать
    user_id = message.from_user.id
    if user_has_uploaded_photos[user_id]:
        keyboard = get_after_photos_keyboard()
    else:
        keyboard = get_after_folder_selection_keyboard()
    
    if current_public_url:
        await message.answer(
            f"📁 <b>Текущая папка</b>\n\n"
            f"Папка: <code>{current_folder}</code>\n\n"
            f"<a href='{current_public_url}'>🔗 Открыть в Яндекс.Диске</a>",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
    else:
        await message.answer(
            f"📁 <b>Текущая папка</b>\n\n"
            f"Папка: <code>{current_folder}</code>\n\n"
            f"❌ Не удалось получить ссылку на папку.",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )

@router.message(F.text == "📂 Основная папка (Фото за МК)")
async def open_base_folder(message: Message, state: FSMContext):
    """Обработчик открытия основной папки"""
    # Определяем, какую клавиатуру показывать
    user_id = message.from_user.id
    if user_has_uploaded_photos[user_id]:
        keyboard = get_after_photos_keyboard()
    else:
        keyboard = get_after_folder_selection_keyboard()
    
    await message.answer(
        f"📂 <b>Основная папка</b>\n\n"
        f"Папка: <code>{BASE_YANDEX_FOLDER}</code>\n\n"
        f"<a href='{BASE_FOLDER_URL}'>🔗 Открыть в Яндекс.Диске</a>",
        reply_markup=keyboard,
        parse_mode=ParseMode.HTML
    )

@router.message(F.text == "📂 Выбрать другую папку")
async def choose_another_folder(message: Message, state: FSMContext):
    """Обработчик выбора другой папки"""
    # Сбрасываем флаг загрузки фото при смене папки
    user_id = message.from_user.id
    user_has_uploaded_photos[user_id] = False
    user_skip_text_requested[user_id] = False
    
    await state.set_state(PhotoStates.choosing_folder)
    await show_folder_contents(message, BASE_YANDEX_FOLDER)

@router.message(F.text == "🏠 В главное меню")
async def back_to_main_menu(message: Message, state: FSMContext):
    """Обработчик возврата в главное меню"""
    # Сбрасываем состояние
    user_id = message.from_user.id
    await state.clear()
    
    # Очищаем временные данные пользователя
    if user_id in user_photo_data:
        # Удаляем временные файлы
        for photo_data in user_photo_data[user_id]:
            temp_file_path = photo_data.get('temp_file_path')
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
        user_photo_data[user_id] = []
    
    if user_id in user_album_data:
        del user_album_data[user_id]
    if user_id in user_batch_data:
        del user_batch_data[user_id]
    if user_id in user_text_data:
        del user_text_data[user_id]
    
    # Сбрасываем флаги
    user_has_uploaded_photos[user_id] = False
    user_skip_text_requested[user_id] = False
    
    await send_welcome(message, state)

@router.message(F.text == "📸 Загрузить фото")
async def start_photo_upload_after_folder(message: Message, state: FSMContext):
    """Начало загрузки фото после выбора папки"""
    if not disk_uploader:
        await message.answer("❌ Яндекс.Диск не подключен.", reply_markup=get_main_keyboard())
        return
    
    user_id = message.from_user.id
    
    # Проверяем, выбрана ли папка
    current_folder = disk_uploader.get_current_folder()
    if current_folder == BASE_YANDEX_FOLDER:
        # Если выбрана основная папка, предлагаем создать свою
        await message.answer(
            "📸 <b>Загрузка фотографий</b>\n\n"
            f"Текущая папка: <code>{current_folder}</code>\n\n"
            "⚠️ <b>Внимание!</b> Вы загружаете фото в основную папку.\n"
            "Рекомендуется создать свою личную папку для организации фото.\n\n"
            "Вы можете:\n"
            "• Продолжить загрузку в основную папку\n"
            "• Создать новую папку (кнопка '➕ Создать папку')\n"
            "• Выбрать другую папку (кнопка '📂 Выбрать другую папку')",
            reply_markup=get_after_folder_selection_keyboard(),
            parse_mode=ParseMode.HTML
        )
        return
    
    user_photo_data[user_id] = []
    user_batch_data[user_id] = {'mode': 'batch'}  # Режим пакетной загрузки
    
    await state.set_state(PhotoStates.processing_album)
    await message.answer(
        "📚 <b>Режим загрузки фотографий</b>\n\n"
        f"📁 Папка для загрузки: <code>{current_folder}</code>\n\n"
        "Отправь мне фотографии - можно выбрать сразу несколько в галерее\n"
        "После отправки я покажу все полученные фото и спрошу, как их загрузить.\n\n"
        "Или используй /cancel для отмены.",
        reply_markup=get_cancel_keyboard(),
        parse_mode=ParseMode.HTML
    )

@router.message(F.text == "📸 Загрузить ещё фото")
async def start_photo_upload_again(message: Message, state: FSMContext):
    """Начало загрузки ещё фото после предыдущей загрузки"""
    user_id = message.from_user.id
    user_skip_text_requested[user_id] = False
    await start_photo_upload_after_folder(message, state)

@router.message(F.text == "⏭️ Пропустить текст (завершить)")
async def skip_text_request(message: Message, state: FSMContext):
    """Запрос на пропуск ввода текста"""
    user_id = message.from_user.id
    
    if not user_has_uploaded_photos[user_id]:
        await message.answer(
            "❌ <b>Нет загруженных фото для завершения</b>\n\n"
            "Сначала загрузите хотя бы одну фотографию.",
            reply_markup=get_after_folder_selection_keyboard(),
            parse_mode=ParseMode.HTML
        )
        return
    
    await state.set_state(TextStates.skip_text_confirmation)
    await message.answer(
        "⚠️ <b>Вы уверены, что хотите пропустить ввод текста?</b>\n\n"
        "В этом случае информация о мастер-классе не будет сохранена в Excel.\n"
        "Будут сохранены только загруженные фотографии.\n\n"
        "Выберите действие:",
        reply_markup=get_skip_text_confirmation_keyboard(),
        parse_mode=ParseMode.HTML
    )

@router.message(TextStates.skip_text_confirmation, F.text == "✅ Да, пропустить")
async def confirm_skip_text(message: Message, state: FSMContext):
    """Подтверждение пропуска текста"""
    user_id = message.from_user.id
    
    # Сохраняем запись о том, что пользователь пропустил текст
    user_skip_text_requested[user_id] = True
    
    # Сбрасываем флаг загрузки фото, чтобы можно было начать заново
    user_has_uploaded_photos[user_id] = False
    
    await state.clear()
    await message.answer(
        "✅ <b>Операция завершена!</b>\n\n"
        "Фотографии успешно загружены.\n"
        "Текст не был сохранен.\n\n"
        "Возвращаемся в главное меню.",
        reply_markup=get_main_keyboard(),
        parse_mode=ParseMode.HTML
    )

@router.message(TextStates.skip_text_confirmation, F.text == "❌ Нет, ввести текст")
async def cancel_skip_text(message: Message, state: FSMContext):
    """Отмена пропуска текста - возврат к вводу"""
    await state.set_state(TextStates.waiting_for_fio)
    await message.answer(
        "📝 <b>Введите ваше ФИО</b>\n\n"
        "Пожалуйста, введи свои Фамилию, Имя и Отчество полностью.\n"
        "Пример: <code>Иванов Иван Иванович</code>\n\n"
        "Это нужно для идентификации в таблице.\n\n"
        "<b>Внимание! Отменить этот процесс нельзя.</b>",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode=ParseMode.HTML
    )

@router.message(F.text == "📝 Отправить текст")
async def start_text_input(message: Message, state: FSMContext):
    """Начало ввода текста (только после загрузки фото)"""
    user_id = message.from_user.id
    
    # Проверяем, загружал ли пользователь фото
    if not user_has_uploaded_photos[user_id]:
        await message.answer(
            "❌ <b>Текст можно отправить только после загрузки фотографий!</b>\n\n"
            "Сначала загрузи хотя бы одну фотографию, а затем возвращайся к отправке текста.",
            reply_markup=get_after_folder_selection_keyboard(),
            parse_mode=ParseMode.HTML
        )
        return
    
    # Убираем клавиатуру - пользователь должен вводить текст
    await state.set_state(TextStates.waiting_for_fio)
    await message.answer(
        "📝 <b>Введите ваше ФИО</b>\n\n"
        "Пожалуйста, введи свои Фамилию, Имя и Отчество полностью.\n"
        "Пример: <code>Иванов Иван Иванович</code>\n\n"
        "Это нужно для идентификации в таблице.\n\n"
        "<b>Внимание! Отменить этот процесс нельзя.</b>",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode=ParseMode.HTML
    )

@router.message(TextStates.waiting_for_fio)
async def process_fio_input(message: Message, state: FSMContext):
    """Обработка ввода ФИО"""
    fio = message.text.strip()
    
    if not fio:
        await message.answer("❌ ФИО не может быть пустым. Попробуйте снова:")
        return
    
    # Проверяем, что ФИО состоит минимум из двух слов
    words = fio.split()
    if len(words) < 2:
        await message.answer(
            "❌ Пожалуйста, введи полное ФИО (минимум фамилия и имя).\n"
            "Пример: <code>Иванов Иван Иванович</code>",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Сохраняем ФИО
    user_id = message.from_user.id
    user_fio_data[user_id] = fio
    
    # Обновляем информацию о пользователе в Excel
    excel_manager.update_user_info(
        user_id,
        message.from_user.username,
        fio
    )
    
    await state.set_state(TextStates.waiting_for_text)
    await message.answer(
        f"✅ ФИО сохранено: <code>{fio}</code>\n\n"
        "📝 <b>Теперь отправь сюда название для сохранения</b>\n"
        "Введи текст, который сохранится в Excel таблице.\n"
        "После ввода текста нужно будет ввести описание МК (обязательно).\n\n"
        "Пример: 'МК 1: Проблема и человек.'\n\n"
        "<b>Отменить нельзя - нужно обязательно ввести текст.</b>",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode=ParseMode.HTML
    )

@router.message(TextStates.waiting_for_text)
async def process_text_input(message: Message, state: FSMContext):
    """Обработка введенного текста"""
    text = message.text.strip()
    
    if not text:
        await message.answer("❌ Текст не может быть пустым. Попробуйте снова:")
        return
    
    # Сохраняем текст во временном хранилище
    user_id = message.from_user.id
    user_text_data[user_id] = {
        'text': text,
        'timestamp': datetime.now()
    }
    
    # Переходим к вводу описания
    await state.set_state(TextStates.waiting_for_text_description)
    await message.answer(
        f"✅ Текст получен:\n\n<code>{text[:100]}{'...' if len(text) > 100 else ''}</code>\n\n"
        "📝 <b>Теперь введи описание МК</b>\n\n"
        "Описание обязательно для сохранения.\n"
        "Например: 'На мастер-классе было разобрано следующее: кто такой предприниматель, как создаётся бизнес, и то, как определить реальность проблемы. Участники хорошо усвоили материал и теперь они готовы общаться со всоей ЦА для выявления проблем'\n\n"
        "<b>После ввода описания данные будут сохранены.</b>",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode=ParseMode.HTML
    )

@router.message(TextStates.waiting_for_text_description)
async def process_text_description(message: Message, state: FSMContext):
    """Обработка введенного описания"""
    description = message.text.strip()
    user_id = message.from_user.id
    
    if not description:
        await message.answer("❌ Описание не может быть пустым. Попробуйте снова:")
        return
    
    if user_id in user_text_data:
        text_data = user_text_data[user_id]
        fio = user_fio_data.get(user_id, "Не указано")
        
        # Сохраняем в Excel с описанием
        success = excel_manager.save_message(
            user_id=user_id,
            fio=fio,
            text=text_data['text'],
            description=description,
            msg_type="text"
        )
        
        if success:
            # Устанавливаем флаг, что пользователь отправил текст
            user_has_sent_text[user_id] = True
            
            await message.answer(
                f"✅ <b>Текст успешно сохранен в Excel!</b>\n\n"
                f"👤 ФИО: <code>{fio}</code>\n"
                f"📝 Текст: <code>{text_data['text'][:100]}{'...' if len(text_data['text']) > 100 else ''}</code>\n"
                f"📌 Описание: <code>{description}</code>\n\n"
                "Возвращаемся в главное меню.",
                reply_markup=get_main_keyboard(),
                parse_mode=ParseMode.HTML
            )
        else:
            await message.answer(
                "❌ Ошибка при сохранении текста.",
                reply_markup=get_main_keyboard()
            )
        
        # Очищаем временные данные
        del user_text_data[user_id]
        # Сбрасываем флаг загрузки фото
        user_has_uploaded_photos[user_id] = False
        await state.clear()
    else:
        await message.answer("❌ Данные не найдены. Начните заново.")
        await state.clear()

@router.message(Command("export"))
@router.message(F.text == "📊 Показать данные")
async def export_data(message: Message, state: FSMContext):
    """Экспорт всех данных в Excel файл"""
    try:
        # Получаем все сообщения
        messages = excel_manager.get_all_messages()
        
        if not messages:
            await message.answer(
                "📊 <b>Нет данных для экспорта</b>\n\n"
                "Пока не было сохранено ни одного сообщения.",
                reply_markup=get_main_keyboard(),
                parse_mode=ParseMode.HTML
            )
            return
        
        # Создаем временную копию для отправки
        import shutil
        temp_file = f"temp_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(EXCEL_FILE, temp_file)
        
        # Отправляем файл
        document = FSInputFile(temp_file)
        await message.answer_document(
            document,
            caption=f"📊 <b>Экспорт всех данных</b>\n\n"
                    f"Всего записей: {len(messages)}\n"
                    f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            parse_mode=ParseMode.HTML
        )
        
        # Удаляем временный файл
        os.remove(temp_file)
        
        # Показываем статистику
        text_count = sum(1 for m in messages if m.get('Тип') == 'text')
        photo_count = sum(1 for m in messages if m.get('Тип') == 'photo')
        
        await message.answer(
            f"📊 <b>Статистика</b>\n\n"
            f"📝 Текстовых сообщений: {text_count}\n"
            f"📸 Фотографий: {photo_count}\n"
            f"👥 Всего пользователей: {len(set(m.get('User ID') for m in messages))}",
            reply_markup=get_main_keyboard(),
            parse_mode=ParseMode.HTML
        )
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных: {e}")
        await message.answer(
            "❌ Ошибка при экспорте данных.",
            reply_markup=get_main_keyboard()
        )

@router.message(Command("stats"))
@router.message(F.text == "📊 Статистика")
async def show_statistics(message: Message, state: FSMContext):
    """Показывает статистику по сохраненным сообщениям"""
    try:
        stats = excel_manager.get_statistics()
        
        if not stats or stats['total_messages'] == 0:
            await message.answer(
                "📊 <b>Статистика</b>\n\n"
                "Пока нет сохраненных сообщений.",
                reply_markup=get_main_keyboard(),
                parse_mode=ParseMode.HTML
            )
            return
        
        # Получаем информацию о бэкапах
        backups = excel_manager.list_backups()
        total_backup_size = sum(b['size'] for b in backups) / (1024 * 1024)
        
        # Формируем сообщение со статистикой
        stat_text = (
            f"📊 <b>Общая статистика</b>\n\n"
            f"📝 Всего сообщений: {stats['total_messages']}\n"
            f"📸 Фотографий: {stats['photo_messages']}\n"
            f"📄 Текстовых записей: {stats['text_messages']}\n"
            f"👥 Уникальных пользователей: {stats['unique_users']}\n\n"
            f"💾 <b>Информация о бэкапах</b>\n"
            f"📁 Всего бэкапов: {len(backups)}\n"
            f"💿 Общий размер: {total_backup_size:.2f} МБ\n"
            f"📅 Автоматические бэкапы: каждые {BACKUP_INTERVAL_HOURS} часов\n"
            f"🗑️ Хранение: {BACKUP_RETENTION_DAYS} дней\n\n"
            f"<b>Детальная статистика по пользователям:</b>\n"
        )
        
        for user in stats['users'][:10]:  # Показываем первых 10 пользователей
            stat_text += (
                f"\n👤 {user['fio']} (ID: {user['user_id']})\n"
                f"  • Сообщений: {user['messages_count']}\n"
                f"  • Первое: {user['first_message']}\n"
                f"  • Последнее: {user['last_message']}\n"
            )
        
        if len(stats['users']) > 10:
            stat_text += f"\n<i>... и еще {len(stats['users']) - 10} пользователей</i>"
        
        # Разбиваем на части если сообщение слишком длинное
        if len(stat_text) > 4000:
            parts = [stat_text[i:i+4000] for i in range(0, len(stat_text), 4000)]
            for i, part in enumerate(parts, 1):
                await message.answer(
                    part + (f"\n\n<i>Часть {i}/{len(parts)}</i>" if len(parts) > 1 else ""),
                    parse_mode=ParseMode.HTML,
                    reply_markup=get_main_keyboard() if i == len(parts) else None
                )
        else:
            await message.answer(stat_text, parse_mode=ParseMode.HTML, reply_markup=get_main_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка при показе статистики: {e}")
        await message.answer("❌ Ошибка при получении статистики.", reply_markup=get_main_keyboard())

@router.message(F.text == "❓ Помощь")
async def help_button(message: Message, state: FSMContext):
    """Обработчик кнопки Помощь"""
    await send_welcome(message, state)

async def show_folder_contents(message: Message, folder_path: str):
    """Показывает содержимое папки"""
    global folder_index_counter
    folder_index_counter = 0  # Сбрасываем счетчик для новой сессии
    folder_paths.clear()  # Очищаем старые пути
    
    if not disk_uploader:
        await message.answer("❌ Яндекс.Диск не подключен", reply_markup=get_main_keyboard())
        return
    
    folders = disk_uploader.list_folders(folder_path)
    is_root = (folder_path == BASE_YANDEX_FOLDER)
    
    # Форматируем список папок для отображения
    folder_list = []
    for f in folders[:10]:
        name = f.get('name', 'Без названия')
        folder_list.append(f"📁 {name[:30]}{'...' if len(name) > 30 else ''}")
    
    folder_text = "\n".join(folder_list)
    if len(folders) > 10:
        folder_text += f"\n... и еще {len(folders) - 10} папок"
    
    await message.answer(
        f"📂 <b>Выбор папки</b>\n\n"
        f"Текущая папка: <code>{folder_path}</code>\n\n"
        f"<b>Доступные папки ({len(folders)}):</b>\n"
        f"<i>(показаны в две колонки ниже)</i>\n\n"
        f"{folder_text}\n\n"
        f"Выберите папку для просмотра содержимого или нажмите '✅ Выбрать эту папку' для загрузки фото:",
        reply_markup=get_folder_keyboard(folders, folder_path, is_root),
        parse_mode=ParseMode.HTML
    )

@router.message(F.text == "📂 Выбрать папку для загрузки")
@router.message(Command("folder"))
async def choose_folder(message: Message, state: FSMContext):
    """Начало выбора папки"""
    await state.set_state(PhotoStates.choosing_folder)
    await show_folder_contents(message, BASE_YANDEX_FOLDER)

@router.message(Command("newfolder"))
@router.message(F.text == "➕ Создать папку")
async def create_new_folder_command(message: Message, state: FSMContext):
    """Создание новой папки из главного меню"""
    if not disk_uploader:
        await message.answer("❌ Яндекс.Диск не подключен", reply_markup=get_main_keyboard())
        return
    
    current_folder = disk_uploader.get_current_folder()
    
    await state.set_state(PhotoStates.waiting_for_folder_name)
    await state.update_data(parent_folder=current_folder, from_command=True)
    await message.answer(
        f"➕ <b>Создание новой папки</b>\n\n"
        f"Текущая папка: <code>{current_folder}</code>\n\n"
        "Введите название для новой папки:\n"
        "(используйте /cancel или кнопку ниже для отмены)",
        reply_markup=get_cancel_keyboard()
    )

# Обработчик для медиагрупп (несколько фото в одном сообщении)
@router.message(PhotoStates.processing_album, F.media_group_id)
async def handle_media_group(message: Message, state: FSMContext, album: List[Message] = None):
    """Обработка медиагруппы (несколько фото в одном сообщении)"""
    try:
        user_id = message.from_user.id
        media_group_id = message.media_group_id
        
        # Сохраняем сообщение во временное хранилище
        media_group_storage[media_group_id].append(message)
        
        # Ждем немного, чтобы собрать все сообщения из медиагруппы
        await asyncio.sleep(0.5)
        
        # Получаем все сообщения из хранилища
        messages = media_group_storage.pop(media_group_id, [])
        
        if not messages:
            return
        
        # Счетчик добавленных фото
        added_count = 0
        
        for msg in messages:
            if msg.photo:
                photo = msg.photo[-1]
                await process_batch_photo(msg, state, photo, user_id)
                added_count += 1
            elif msg.document and msg.document.mime_type and 'image' in msg.document.mime_type:
                await process_batch_photo(msg, state, msg.document, user_id)
                added_count += 1
        
        # Сообщаем о количестве добавленных фото
        if added_count > 0:
            await message.answer(
                f"✅ Получено {added_count} фотографий!\n"
                f"Всего в буфере: {len(user_photo_data[user_id])} фото\n\n"
                "Продолжайте отправлять фото или выберите действие ниже.",
                reply_markup=get_batch_upload_keyboard()
            )
    
    except Exception as e:
        logger.error(f"Error handling media group: {e}")
        await message.answer("❌ Ошибка при обработке группы фото.")

# Обработчик для одиночных фото
@router.message(PhotoStates.processing_album, F.photo | F.document)
async def handle_album_photo(message: Message, state: FSMContext):
    """Обработка одиночного фото в режиме альбома"""
    try:
        user_id = message.from_user.id
        
        # Проверяем, не является ли это частью медиагруппы
        if message.media_group_id:
            # Если это часть медиагруппы, игнорируем - будет обработано в handle_media_group
            return
        
        if message.photo:
            photo = message.photo[-1]
            await process_batch_photo(message, state, photo, user_id)
        elif message.document and message.document.mime_type and 'image' in message.document.mime_type:
            await process_batch_photo(message, state, message.document, user_id)
        else:
            await message.answer("❌ Пожалуйста, отправьте фото.", reply_markup=get_batch_upload_keyboard())
            return
        
        await message.answer(
            f"✅ Фотография добавлена.\n"
            f"Всего в буфере: {len(user_photo_data[user_id])} фото\n\n"
            "Отправьте следующую фотографию или выберите действие ниже.",
            reply_markup=get_batch_upload_keyboard()
        )
    
    except Exception as e:
        logger.error(f"Error handling album photo: {e}")
        await message.answer("❌ Ошибка при обработке фото.")

async def process_batch_photo(message: Message, state: FSMContext, photo, user_id: int):
    """Обработка фото для пакетной загрузки"""
    try:
        if isinstance(photo, types.PhotoSize):
            file_info = await bot.get_file(photo.file_id)
            original_ext = ".jpg"
        else:
            file_info = await bot.get_file(photo.file_id)
            if photo.file_name:
                _, original_ext = os.path.splitext(photo.file_name)
                if not original_ext:
                    original_ext = ".jpg"
            else:
                original_ext = ".jpg"
        
        downloaded_file = await bot.download_file(file_info.file_path)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:19]
        temp_name = f"temp_batch_{user_id}_{len(user_photo_data[user_id])}_{timestamp}{original_ext}"
        
        temp_file_path = os.path.join(TEMP_FOLDER, temp_name)
        with open(temp_file_path, 'wb') as f:
            f.write(downloaded_file.getvalue())
        
        photo_data = {
            'temp_file_path': temp_file_path,
            'original_ext': original_ext,
            'message_id': message.message_id,
            'auto_name': f"photo_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{len(user_photo_data[user_id])+1}{original_ext}"
        }
        
        user_photo_data[user_id].append(photo_data)
        
    except Exception as e:
        logger.error(f"Error processing batch photo: {e}")
        raise e

@router.message(PhotoStates.processing_album, F.text == "📝 Дать название каждому")
async def start_naming_batch(message: Message, state: FSMContext):
    """Начало процесса именования каждого фото"""
    user_id = message.from_user.id
    
    if user_id not in user_photo_data or not user_photo_data[user_id]:
        await message.answer("❌ Нет фото для обработки. Сначала отправьте фотографии.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    total_photos = len(user_photo_data[user_id])
    user_album_data[user_id] = {
        'total': total_photos,
        'processed': 0,
        'mode': 'naming'
    }
    
    await state.set_state(PhotoStates.waiting_for_custom_name)
    photo_data = user_photo_data[user_id][0]
    
    await message.answer(
        f"📸 <b>Начинаем именование {total_photos} фото</b>\n\n"
        f"Фото 1 из {total_photos}\n\n"
        "Как вы хотите назвать этот файл?\n\n"
        "<b>📝 Дать своё название</b> - Вы вводите своё название файла\n"
        "<b>🔄 Использовать автоназвание</b> - Бот сам сгенерирует название\n\n"
        f"Автоназвание: {photo_data['auto_name']}",
        reply_markup=get_upload_options_keyboard()
    )

@router.message(PhotoStates.processing_album, F.text == "🔄 Все с автоназваниями")
async def upload_all_batch_with_autonames(message: Message, state: FSMContext):
    """Загрузка всех фото с автоназваниями"""
    user_id = message.from_user.id
    
    if user_id not in user_photo_data or not user_photo_data[user_id]:
        await message.answer("❌ Нет фото для загрузки.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    total_photos = len(user_photo_data[user_id])
    status_msg = await message.answer(f"⏳ Начинаю загрузку {total_photos} фото с автоназваниями...")
    
    success_count = 0
    failed_count = 0
    failed_files = []
    
    while user_photo_data[user_id]:
        photo_data = user_photo_data[user_id][0]
        temp_file_path = photo_data['temp_file_path']
        auto_name = photo_data['auto_name']
        
        drive_link = disk_uploader.upload_file(temp_file_path, auto_name)
        
        try:
            os.remove(temp_file_path)
        except:
            pass
        
        if drive_link:
            success_count += 1
        else:
            failed_count += 1
            failed_files.append(auto_name)
        
        user_photo_data[user_id].pop(0)
    
    # Устанавливаем флаг, что пользователь загрузил фото
    user_has_uploaded_photos[user_id] = True
    
    current_folder = disk_uploader.get_current_folder()
    
    # Получаем ссылки на текущую и основную папки
    current_public_url = disk_uploader.get_public_url(current_folder)
    
    result_text = (
        f"📊 <b>Загрузка завершена!</b>\n\n"
        f"📁 Текущая папка: <code>{current_folder}</code>\n"
        f"✅ Успешно: {success_count} фото\n"
    )
    
    if failed_count > 0:
        result_text += f"❌ Ошибок: {failed_count} фото\n"
        if failed_files:
            result_text += f"Не удалось загрузить: {', '.join(failed_files[:3])}\n"
    
    result_text += f"\n📂 <a href='{BASE_FOLDER_URL}'>Основная папка (Фото за МК)</a>\n\n"
    result_text += "Что делаем дальше?"
    
    await message.answer(
        result_text,
        reply_markup=get_after_photos_keyboard(),  # Обновленная клавиатура с кнопкой пропуска
        parse_mode=ParseMode.HTML
    )
    
    if user_id in user_batch_data:
        del user_batch_data[user_id]
    
    await state.clear()

@router.message(PhotoStates.waiting_for_custom_name)
async def process_custom_name(message: Message, state: FSMContext):
    """Обработка ввода пользовательского названия файла"""
    user_id = message.from_user.id
    
    if user_id not in user_photo_data or not user_photo_data[user_id]:
        await message.answer("❌ Данные о фото не найдены. Начните заново.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    custom_name = message.text.strip()
    
    if not custom_name:
        await message.answer("❌ Название файла не может быть пустым. Попробуйте снова:", reply_markup=get_cancel_keyboard())
        return
    
    invalid_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    if any(char in custom_name for char in invalid_chars):
        await message.answer("❌ Название файла содержит недопустимые символы. Попробуйте снова:", reply_markup=get_cancel_keyboard())
        return
    
    photo_data = user_photo_data[user_id][0]
    temp_file_path = photo_data['temp_file_path']
    original_ext = photo_data['original_ext']
    
    if '.' not in custom_name:
        custom_name = f"{custom_name}{original_ext}"
    elif not custom_name.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
        name_without_ext = os.path.splitext(custom_name)[0]
        custom_name = f"{name_without_ext}{original_ext}"
    
    photo_data['auto_name'] = custom_name
    
    await process_file_upload(message, state, temp_file_path, custom_name, photo_data)
    
    user_photo_data[user_id].pop(0)
    
    if user_photo_data[user_id]:
        if user_id in user_album_data:
            user_album_data[user_id]['processed'] += 1
            current_index = user_album_data[user_id]['processed'] + 1
            total = user_album_data[user_id]['total']
            
        await state.set_state(PhotoStates.waiting_for_custom_name)
        await message.answer(
            f"📸 <b>Следующее фото получено!</b>\n\n"
            f"Фото {current_index} из {total}\n\n"
            "Как вы хотите назвать этот файл?\n\n"
            "<b>📝 Дать своё название</b> - Вы вводите своё название файла\n"
            "<b>🔄 Использовать автоназвание</b> - Бот сам сгенерирует название\n\n"
            "Автоназвание: " + user_photo_data[user_id][0]['auto_name'],
            reply_markup=get_upload_options_keyboard()
        )
    else:
        if user_id in user_album_data:
            total_processed = user_album_data[user_id].get('processed', 0) + 1
            total = user_album_data[user_id].get('total', 0)
            
            # Устанавливаем флаг, что пользователь загрузил фото
            user_has_uploaded_photos[user_id] = True
            
            # Получаем ссылки на текущую и основную папки
            current_folder = disk_uploader.get_current_folder()
            current_public_url = disk_uploader.get_public_url(current_folder)
            
            result_text = (
                f"🎉 <b>Фото обработаны</b>\n\n"
                f"📊 Загружено фото: {total_processed}/{total}\n"
                f"📁 Текущая папка: <code>{current_folder}</code>\n\n"
            )
            
            result_text += f"📂 <a href='{BASE_FOLDER_URL}'>Основная папка (Фото за МК)</a>\n\n"
            result_text += "Что делаем дальше?"
            
            await message.answer(
                result_text,
                reply_markup=get_after_photos_keyboard(),  # Обновленная клавиатура с кнопкой пропуска
                parse_mode=ParseMode.HTML
            )
            
            del user_album_data[user_id]
        
        if user_id in user_batch_data:
            del user_batch_data[user_id]
        
        await state.clear()

@router.message(PhotoStates.waiting_for_custom_name, F.text == "📝 Дать своё название")
async def request_custom_name(message: Message):
    """Запрос пользовательского названия"""
    await message.answer(
        "📝 <b>Введите название для файла:</b>\n\n"
        "Примеры:\n"
        "• Мое фото\n"
        "• Отпуск 2024\n"
        "• Концерт группа\n\n"
        "Расширение файла будет добавлено автоматически.",
        reply_markup=get_cancel_keyboard()
    )

@router.message(PhotoStates.waiting_for_custom_name, F.text == "🔄 Использовать автоназвание")
async def use_auto_name(message: Message, state: FSMContext):
    """Использование автосгенерированного названия"""
    user_id = message.from_user.id
    
    if user_id not in user_photo_data or not user_photo_data[user_id]:
        await message.answer("❌ Данные о фото не найдены. Начните заново.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    photo_data = user_photo_data[user_id][0]
    temp_file_path = photo_data['temp_file_path']
    auto_name = photo_data['auto_name']
    
    await process_file_upload(message, state, temp_file_path, auto_name, photo_data)
    
    user_photo_data[user_id].pop(0)
    
    if user_photo_data[user_id]:
        if user_id in user_album_data:
            user_album_data[user_id]['processed'] += 1
            current_index = user_album_data[user_id]['processed'] + 1
            total = user_album_data[user_id]['total']
            
        await state.set_state(PhotoStates.waiting_for_custom_name)
        await message.answer(
            f"📸 <b>Следующее фото получено!</b>\n\n"
            f"Фото {current_index} из {total}\n\n"
            "Как вы хотите назвать этот файл?\n\n"
            "<b>📝 Дать своё название</b> - Вы вводите своё название файла\n"
            "<b>🔄 Использовать автоназвание</b> - Бот сам сгенерирует название\n\n"
            "Автоназвание: " + user_photo_data[user_id][0]['auto_name'],
            reply_markup=get_upload_options_keyboard()
        )
    else:
        if user_id in user_album_data:
            total_processed = user_album_data[user_id].get('processed', 0) + 1
            total = user_album_data[user_id].get('total', 0)
            
            # Устанавливаем флаг, что пользователь загрузил фото
            user_has_uploaded_photos[user_id] = True
            
            # Получаем ссылки на текущую и основную папки
            current_folder = disk_uploader.get_current_folder()
            current_public_url = disk_uploader.get_public_url(current_folder)
            
            result_text = (
                f"🎉 <b>Фото обработаны</b>\n\n"
                f"📊 Загружено фото: {total_processed}/{total}\n"
                f"📁 Текущая папка: <code>{current_folder}</code>\n\n"
            )
            
            result_text += f"📂 <a href='{BASE_FOLDER_URL}'>Основная папка (Фото за МК)</a>\n\n"
            result_text += "Что делаем дальше?"
            
            await message.answer(
                result_text,
                reply_markup=get_after_photos_keyboard(),  # Обновленная клавиатура с кнопкой пропуска
                parse_mode=ParseMode.HTML
            )
            
            del user_album_data[user_id]
        
        if user_id in user_batch_data:
            del user_batch_data[user_id]
        
        await state.clear()

@router.message(PhotoStates.waiting_for_custom_name, F.text == "📁 Загрузить все с автоназваниями")
async def upload_all_with_autonames(message: Message, state: FSMContext):
    """Загрузка всех оставшихся фото с автоназваниями"""
    user_id = message.from_user.id
    
    if user_id not in user_photo_data or not user_photo_data[user_id]:
        await message.answer("❌ Нет фото для загрузки.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    total_photos = len(user_photo_data[user_id])
    await message.answer(f"⏳ Начинаю загрузку {total_photos} фото с автоназваниями...")
    
    success_count = 0
    failed_count = 0
    
    while user_photo_data[user_id]:
        photo_data = user_photo_data[user_id][0]
        temp_file_path = photo_data['temp_file_path']
        auto_name = photo_data['auto_name']
        
        drive_link = disk_uploader.upload_file(temp_file_path, auto_name)
        
        try:
            os.remove(temp_file_path)
        except:
            pass
        
        if drive_link:
            success_count += 1
        else:
            failed_count += 1
        
        user_photo_data[user_id].pop(0)
    
    # Устанавливаем флаг, что пользователь загрузил фото
    user_has_uploaded_photos[user_id] = True
    
    current_folder = disk_uploader.get_current_folder()
    
    # Получаем ссылки на текущую и основную папки
    current_public_url = disk_uploader.get_public_url(current_folder)
    
    result_text = (
        f"📊 <b>Загрузка завершена!</b>\n\n"
        f"📁 Текущая папка: <code>{current_folder}</code>\n"
        f"✅ Успешно: {success_count} фото\n"
    )
    
    if failed_count > 0:
        result_text += f"❌ Ошибок: {failed_count} фото\n"
    
    result_text += f"\n📂 <a href='{BASE_FOLDER_URL}'>Основная папка (Фото за МК)</a>\n\n"
    
    if user_id in user_album_data:
        total = user_album_data[user_id].get('total', 0)
        result_text += f"\n📚 {total} фото полностью обработаны!"
        del user_album_data[user_id]
    
    await message.answer(
        result_text,
        reply_markup=get_after_photos_keyboard(),  # Обновленная клавиатура с кнопкой пропуска
        parse_mode=ParseMode.HTML
    )
    
    if user_id in user_batch_data:
        del user_batch_data[user_id]
    
    await state.clear()

async def process_file_upload(message: Message, state: FSMContext, temp_file_path: str, file_name: str, photo_data: dict):
    """Обработка загрузки файла"""
    try:
        user_id = message.from_user.id
        
        status_msg = await message.answer("⏳ Загружаю файл в Яндекс.Диск...")
        drive_link = disk_uploader.upload_file(temp_file_path, file_name)
        
        try:
            os.remove(temp_file_path)
        except:
            pass
        
        if drive_link:
            current_folder = disk_uploader.get_current_folder()
            await status_msg.edit_text(
                f"✅ <b>Файл успешно загружен!</b>\n\n"
                f"📁 Имя файла: <code>{file_name}</code>\n"
                f"📂 Папка: <code>{current_folder}</code>\n"
                f"🔗 <a href='{drive_link}'>Ссылка на файл</a>"
            )
        else:
            await status_msg.edit_text("❌ Ошибка при загрузке в Яндекс.Диск")
    
    except Exception as e:
        logger.error(f"Error processing file upload: {e}")
        await message.answer("❌ Произошла ошибка при загрузке файла.", reply_markup=get_main_keyboard())
        await state.clear()

@router.message(Command("cancel"))
@router.message(F.text == "❌ Отмена")
async def cancel_upload(message: Message, state: FSMContext):
    """Отмена текущей операции"""
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("❌ Нет активной операции для отмены.", reply_markup=get_main_keyboard())
        return
    
    # Проверяем, не находится ли пользователь в процессе ввода текста
    if current_state in [TextStates.waiting_for_fio.state, TextStates.waiting_for_text.state, TextStates.waiting_for_text_description.state]:
        await message.answer(
            "❌ <b>Нельзя отменить процесс после начала ввода текста!</b>\n\n"
            "Вы ОБЯЗАНЫ ввести ФИО, текст и описание.\n"
            "Пожалуйста, продолжайте ввод.",
            reply_markup=ReplyKeyboardRemove()
        )
        return
    
    user_id = message.from_user.id
    if user_id in user_photo_data:
        for photo_data in user_photo_data[user_id]:
            temp_file_path = photo_data.get('temp_file_path')
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
        user_photo_data[user_id] = []
    
    if user_id in user_album_data:
        del user_album_data[user_id]
    
    if user_id in user_batch_data:
        del user_batch_data[user_id]
    
    if user_id in user_text_data:
        del user_text_data[user_id]
    
    await state.clear()
    await message.answer("❌ Операция отменена.", reply_markup=get_main_keyboard())

@router.message(PhotoStates.waiting_for_folder_name)
async def process_folder_name(message: Message, state: FSMContext):
    """Обработка ввода имени новой папки"""
    folder_name = message.text.strip()
    
    if not folder_name:
        await message.answer("❌ Имя папки не может быть пустым. Попробуйте снова:", reply_markup=get_cancel_keyboard())
        return
    
    invalid_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    if any(char in folder_name for char in invalid_chars):
        await message.answer("❌ Имя папки содержит недопустимые символы. Попробуйте снова:", reply_markup=get_cancel_keyboard())
        return
    
    data = await state.get_data()
    parent_folder = data.get('parent_folder', disk_uploader.get_current_folder())
    from_command = data.get('from_command', False)
    
    new_folder_path = disk_uploader.create_folder(folder_name, parent_folder)
    
    if new_folder_path:
        if from_command:
            global folder_counter
            folder_index = folder_counter
            folder_counter += 1
            
            await message.answer(
                f"✅ Папка <code>{folder_name}</code> успешно создана!\n\n"
                f"📂 Расположение: <code>{new_folder_path}</code>\n\n"
                "Хотите перейти в созданную папку?",
                reply_markup=get_folder_creation_keyboard(new_folder_path, folder_index)
            )
        else:
            await message.answer(
                f"✅ Папка <code>{folder_name}</code> успешно создана в <code>{parent_folder}</code>\n\n"
                "Теперь вы можете выбрать её в меню выбора папок.",
                reply_markup=get_main_keyboard()
            )
    else:
        await message.answer(f"❌ Не удалось создать папка. Возможно, она уже существует или произошла ошибка.", reply_markup=get_main_keyboard())
    
    await state.clear()

@router.message(PhotoStates.processing_album)
async def handle_other_messages_in_album_state(message: Message):
    """Обработка других сообщений в режиме альбома"""
    if message.text and message.text not in ["📝 Дать название каждому", "🔄 Все с автоназваниями", "❌ Отмена"]:
        await message.answer(
            "📚 <b>Режим загрузки фотографий</b>\n\n"
            "Пожалуйста, отправьте фотографии или выберите действие из меню ниже.",
            reply_markup=get_batch_upload_keyboard()
        )

@router.message(F.photo | F.document)
async def handle_photo_without_state(message: Message):
    """Обработка фото, отправленных без выбора папки"""
    user_id = message.from_user.id
    
    await message.answer(
        "📤 <b>Чтобы загрузить фото, сначала выберите папку!</b>\n\n"
        "Нажмите кнопку '📂 Выбрать папку для загрузки' в главном меню.\n\n"
        "Используйте /start или кнопку '🏠 Стартуем' для возврата в главное меню.",
        reply_markup=get_main_keyboard()
    )

# ========== CALLBACK HANDLERS ==========
@router.callback_query(lambda c: c.data.startswith('nav_folder:'))
async def navigate_folder_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик навигации по папкам"""
    try:
        folder_index = int(callback_query.data.split(':')[1])
        folder_path = folder_paths.get(folder_index)
        
        if folder_path:
            # Показываем содержимое выбранной папки
            await show_folder_contents(callback_query.message, folder_path)
            await callback_query.answer()
            await callback_query.message.delete()
        else:
            await callback_query.answer("❌ Папка не найдена")
    except Exception as e:
        logger.error(f"Error in navigate_folder_callback: {e}")
        await callback_query.answer("❌ Ошибка при навигации")

@router.callback_query(lambda c: c.data.startswith('select_folder:'))
async def select_folder_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик выбора папки для загрузки"""
    try:
        folder_index = int(callback_query.data.split(':')[1])
        folder_path = folder_paths.get(folder_index)
        
        if folder_path:
            disk_uploader.set_current_folder(folder_path)
            await state.clear()
            
            await callback_query.message.edit_text(
                f"✅ Папка выбрана:\n<code>{folder_path}</code>\n\n"
                "Теперь вы можете загружать фотографии."
            )
            await callback_query.answer()
            await callback_query.message.answer(
                "Что делаем дальше?",
                reply_markup=get_after_folder_selection_keyboard()
            )
        else:
            await callback_query.answer("❌ Папка не найдена")
    except Exception as e:
        logger.error(f"Error in select_folder_callback: {e}")
        await callback_query.answer("❌ Ошибка при выборе папки")

@router.callback_query(lambda c: c.data.startswith('select_new_folder:'))
async def select_new_folder_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик выбора вновь созданной папки"""
    try:
        folder_index = int(callback_query.data.split(':')[1])
        
        # Получаем путь из временного хранилища
        folder_path = folder_paths.get(folder_index)
        
        if folder_path:
            disk_uploader.set_current_folder(folder_path)
            # Очищаем временное хранилище
            if folder_index in folder_paths:
                del folder_paths[folder_index]
            
            await state.clear()
            
            await callback_query.message.edit_text(
                f"✅ Папка выбрана:\n<code>{folder_path}</code>\n\n"
                "Теперь вы можете загружать фотографии."
            )
            await callback_query.answer()
            await callback_query.message.answer(
                "Что делаем дальше?",
                reply_markup=get_after_folder_selection_keyboard()
            )
        else:
            await callback_query.answer("❌ Папка не найдена")
    except Exception as e:
        logger.error(f"Error in select_new_folder_callback: {e}")
        await callback_query.answer("❌ Ошибка при выборе папки")

@router.callback_query(lambda c: c.data == 'create_new_folder')
async def create_folder_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик создания новой папки"""
    current_folder = disk_uploader.get_current_folder()
    
    await state.set_state(PhotoStates.waiting_for_folder_name)
    await state.update_data(parent_folder=current_folder, from_command=False)
    await callback_query.message.edit_text(
        f"➕ <b>Создание новой папки</b>\n\n"
        f"Текущая папка: <code>{current_folder}</code>\n\n"
        "Введите название для новой папки:\n"
        "(используйте /cancel для отмены)"
    )
    await callback_query.answer()

@router.callback_query(lambda c: c.data == 'cancel_folder_selection')
async def cancel_folder_selection_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик отмены выбора папки"""
    await state.clear()
    await callback_query.message.edit_text("❌ Выбор папки отменен.")
    await callback_query.answer()
    await callback_query.message.answer("Что делаем дальше?", reply_markup=get_main_keyboard())

@router.callback_query(lambda c: c.data == 'back_to_main_menu')
async def back_to_main_menu_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик возврата в главное меню по кнопке Назад"""
    await state.clear()
    await callback_query.message.delete()
    await send_welcome(callback_query.message, state)
    await callback_query.answer()

@router.callback_query(lambda c: c.data == 'select_another_folder')
async def select_another_folder_callback(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик выбора другой папки"""
    await callback_query.message.delete()
    await show_folder_contents(callback_query.message, BASE_YANDEX_FOLDER)

# ========== ЗАПУСК БОТА ==========
async def main():
    if not disk_uploader:
        logger.error("Бот не может быть запущен: Яндекс.Диск не подключен")
        return
    
    # Запускаем планировщик бэкапов
    asyncio.create_task(backup_scheduler())
    logger.info("Планировщик бэкапов запущен")
    
    # Очищаем временную папку при запуске
    try:
        for file in os.listdir(TEMP_FOLDER):
            file_path = os.path.join(TEMP_FOLDER, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
        logger.info("Временная папка очищена")
    except Exception as e:
        logger.error(f"Ошибка при очистке временной папки: {e}")
    
    logger.info("Бот запускается...")
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
